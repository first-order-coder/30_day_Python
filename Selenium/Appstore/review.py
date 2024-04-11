from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook,load_workbook
import pandas as pd

option = webdriver.ChromeOptions()
Excel_Name = input('Enter Excel Sheet Name::')
link = input("Enter Webpage Link:")
option.add_argument("start-maximized")
option.add_argument("--disable-notifications")

driver = webdriver.Chrome()

driver.get(link)

action = ActionChains(driver)
action.key_down(Keys.PAGE_DOWN).perform()
action.key_down(Keys.PAGE_DOWN).perform()
action.key_down(Keys.PAGE_DOWN).perform()
action.key_down(Keys.PAGE_DOWN).perform()
time.sleep(3)

see = driver.find_element(By.XPATH, "//span[normalize-space()='See all reviews']")
see.click()

time.sleep(3)

# scroll = driver.find_element(By.XPATH, "//div[@class='odk6He']")
popup = driver.find_element(By.XPATH, "//body/div[@class='VfPpkd-Sx9Kwc cC1eCc UDxLd PzCPDd HQdjr VfPpkd-Sx9Kwc-OWXEXe-FNFY6c']/div[@class='VfPpkd-wzTsW']/div[@role='dialog']/div[@class='VfPpkd-cnG4Wd']/div/div[@class='jgIq1']/div[@class='fysCi']/div[@class='odk6He']/div[1]")
popup.click()

time.sleep(3)

action = ActionChains(driver)
time_out = time.time() + 180
while time.time() < time_out:
    action.key_down(Keys.END).perform()

wb = load_workbook(Excel_Name)
wb = Workbook()
worksheet = wb.active

reviews = driver.find_elements(By.CSS_SELECTOR, "div.RHo1pe") 
for item in reviews:
    item1 = item.text.split("\n")
    item1.remove(item1[1])
    # print(item1)

    worksheet.append(item1)
    wb.save(Excel_Name)

time.sleep(1)

# names = driver.find_elements(By.XPATH, "//div[@class='RHo1pe']//div[@class='X5PpBb']")
# rating = driver.find_elements(By.XPATH, "//body/div[@class='VfPpkd-Sx9Kwc cC1eCc UDxLd PzCPDd HQdjr VfPpkd-Sx9Kwc-OWXEXe-FNFY6c']/div[@class='VfPpkd-wzTsW']/div[@role='dialog']/div[@class='VfPpkd-cnG4Wd']/div/div[@class='jgIq1']/div[@class='fysCi']/div[@class='odk6He']/div/div[2]/header[1]/div[2]/div[1]")
# dates = driver.find_elements(By.CSS_SELECTOR, "span.bp9Aid")
# for name in names:
#     print(name.text)

# for date in dates:
#     print(date.text)

# for review in reviews:
#     print(review.text)
#     print("----------------------------")https://play.google.com/store/apps/details?id=com.google.android.apps.subscriptions.red&hl=en&gl=US

