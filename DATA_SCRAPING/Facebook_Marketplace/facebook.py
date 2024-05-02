from selenium import webdriver
from selenium.webdriver import chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.chrome.options import Options
from turtle import heading
from openpyxl import Workbook, load_workbook

option = Options()

option.add_argument('start-maximized')
option.add_argument("--disable-notifications")

driver = webdriver.Chrome()
driver.get('https://www.facebook.com/marketplace/nyc/search/?query=apartment&exact=false')
time.sleep(5)
gmail = input('Enter Facebook Account Gmail')
name = driver.find_element(By.NAME, 'email')
name.send_keys(gmail)

facebook_password = input('Enter Facebook Password: ')
password = driver.find_element(By.NAME, 'pass')
password.send_keys(facebook_password)

login = driver.find_element(By.CSS_SELECTOR, 'div.x1i10hfl.x1qjc9v5.xjqpnuy.xa49m3k.xqeqjp1.x2hbi6w.x13fuv20.xu3j5b3.x1q0q8m5.x26u7qi.x972fbf.xcfux6l.x1qhh985.xm0m39n.x9f619.x1ypdohk.xdl72j9.x2lah0s.xe8uvvx.xdj266r.x11i5rnm.xat24cr.x1mh8g0r.x2lwn1j.xeuugli.x1n2onr6.x16tdsg8.x1hl2dhg.xggy1nq.x1ja2u2z.x1t137rt.x3nfvp2.x1q0g3np.x87ps6o.x1lku1pv.x1a2a7pz.xtvsq51.xhk9q7s.x1otrzb0.x1i1ezom.x1o6z2jb.x1vqgdyp.x6ikm8r.x10wlt62.xexx8yu.xn6708d.x1120s5i.x1ye3gou')
login.click()

time.sleep(3)

# filter = driver.find_element(By.XPATH, '//*[@id="seo_filters"]/div[2]/div[3]/div[2]/span[2]/label/input')
# filter.send_keys('€ 1,000')
# filter.send_keys(Keys.RETURN)

# time.sleep(10)

action = ActionChains(driver)
timeout = time.time() + 180
while time.time() < timeout:
    action.key_down(Keys.END).perform()

total = driver.find_elements(By.CSS_SELECTOR, 'div.x9f619.x78zum5.xdt5ytf.x1qughib.x1rdy4ex.xz9dl7a.xsag5q8.xh8yej3.xp0eagm.x1nrcals')
info = driver.find_elements(By.CSS_SELECTOR, "span.x1lliihq.x6ikm8r.x10wlt62.x1n2onr6")
prices = driver.find_elements(By.CSS_SELECTOR, 'span.x193iq5w.xeuugli.x13faqbe.x1vvkbs.x1xmvt09.x1lliihq.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.xudqn12.x676frb.x1lkfr7t.x1lbecb7.x1s688f.xzsf02u')
# for price in prices:
#     print(price.text)
#     print("------------------")

# for inf in info:
#     print(inf.text)
#     print("---------------")

wb = load_workbook("Apartments2.xlsx")
wb = Workbook()
ws = wb.active
ws['A1'].value = "Price"
ws['B1'].value = "Description"
ws['C1'].value = "Location"

for item in total:
    item1 = item.text.split("\n")
    price = item1[0]
    info = item1[1]
    Location = item1[2]

    ws.append(item1)
    wb.save('Apartments2.xlsx')

    
    # print(f'Price:{price} info:{info} Location:{Location}')
    # print("----------------------------------------")




    


